"""Document parsers for RAG ingestion."""

import csv
import logging
from pathlib import Path
from typing import Iterator

logger = logging.getLogger(__name__)
SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv", ".docx", ".txt", ".md", ""}


def _parse_txt(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse plain text or markdown."""
    text = path.read_text(encoding="utf-8", errors="replace")
    if text.strip():
        yield text.strip(), {"source": str(path.name), "page": None}


def _parse_csv(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse CSV file."""
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return
    lines = []
    for i, row in enumerate(rows):
        lines.append(" | ".join(str(c) for c in row))
    text = "\n".join(lines)
    if text.strip():
        yield text.strip(), {"source": str(path.name), "page": None}


def _parse_pdf(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse PDF using pypdf."""
    from pypdf import PdfReader

    reader = PdfReader(path)
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            yield text.strip(), {"source": str(path.name), "page": i + 1}


def _parse_excel(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse Excel using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            text = "\n".join(rows)
            yield text.strip(), {"source": f"{path.name} ({sheet_name})", "page": None}
    wb.close()


def _parse_docx(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse Word document using python-docx."""
    from docx import Document

    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    if paragraphs:
        text = "\n\n".join(paragraphs)
        yield text.strip(), {"source": str(path.name), "page": None}


def parse_file(path: Path) -> Iterator[tuple[str, dict]]:
    """Parse a file and yield (text, metadata) tuples."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".txt", ".md") or suffix == "":
        yield from _parse_txt(path)
    elif suffix == ".csv":
        yield from _parse_csv(path)
    elif suffix == ".pdf":
        yield from _parse_pdf(path)
    elif suffix in (".xlsx", ".xls"):
        yield from _parse_excel(path)
    elif suffix == ".docx":
        yield from _parse_docx(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def parse_folder(folder: Path) -> Iterator[tuple[str, dict]]:
    """Parse all supported files in a folder."""
    folder = Path(folder)
    if not folder.exists():
        return
    for path in sorted(folder.iterdir()):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                yield from parse_file(path)
            except Exception as e:
                logger.debug("Parse failed for %s: %s", path.name, e)
